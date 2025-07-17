import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill

HIGHLIGHT_COLOR = "FFFF00"

def get_highlight_matrix(excel_file, sheet_name=None):
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    matrix = []
    for row in ws.iter_rows():
        # Ignore the first column
        matrix.append([
            1 if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb != "00000000" and cell.fill.fill_type else 0
            for cell in row[1:]
        ])
    return matrix



def write_missed_identified_fp_excel(template_file, missed_matrix, identified_matrix, fp_matrix, result_f):
    wb = openpyxl.load_workbook(template_file)
    ws = wb.active
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for i, (missed_row, identified_row, fp_row) in enumerate(zip(missed_matrix, identified_matrix, fp_matrix)):
        for j, (missed, identified, fp) in enumerate(zip(missed_row, identified_row, fp_row)):
            cell = ws.cell(row=i+1, column=j+2)  # +2 to skip first column
            if missed:
                cell.fill = red_fill
            elif identified:
                cell.fill = green_fill
            elif fp:
                cell.fill = yellow_fill
    wb.save(result_f)


def create_error_excel_combined(input_f, output_f, result_f):
    gt = get_highlight_matrix(input_f)
    model = get_highlight_matrix(output_f)
    gt_arr = pd.DataFrame(gt)
    model_arr = pd.DataFrame(model)
    missed = ((gt_arr == 1) & (model_arr == 0)).astype(int).values.tolist()
    identified = ((gt_arr == 1) & (model_arr == 1)).astype(int).values.tolist()
    false_positive = ((gt_arr == 0) & (model_arr == 1)).astype(int).values.tolist()
    write_missed_identified_fp_excel(input_f, missed, identified, false_positive, result_f)
    print(f"Missed (red), identified (green), and overpredicted (yellow) errors saved to {result_f}")

if __name__ == "__main__":
    # Example usage
    create_error_excel_combined(
        input_f="Train.xlsx",
        output_f="Train_output.xlsx",
        result_f="missed_and_identified.xlsx"
    )
