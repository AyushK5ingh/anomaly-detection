# Anomaly Detection Script
# This script performs anomaly detection on an Excel (.xlsx) file
# using both rule-based and statistical (Isolation Forest) methods.
# It highlights detected anomalies in the output Excel file and
# provides detailed error analysis, including confusion matrix, F1 score,
# accuracy, precision, and recall for each column and for all columns combined.
# It also generates a separate file with missed (red), identified (green), and overpredicted (yellow) cells.

import pandas as pd
import numpy as np
from sklearn.ensemble import IsolationForest
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import accuracy
import errors

INPUT= "Train"
INPUT_FILE = INPUT + ".xlsx"
OUTPUT_FILE = INPUT + "_output.xlsx"

# If your input file is in a different directory, set INPUT_FILE to the full path:
# Example:
# INPUT = "Train"
# INPUT_FILE = r"D:/your_folder/Train.xlsx"
# Otherwise, by default, INPUT_FILE = INPUT + ".xlsx" (in the current directory).
HIGHLIGHT_COLOR = "FFFF00"

def infer_column_types(df, threshold=0.8):
    """
    Infer column types for a DataFrame: numeric, datetime, categorical, or mixed.

    Args:
        df (pd.DataFrame): Input data.
        threshold (float): Proportion threshold for type assignment.

    Returns:
        dict: Mapping of column name to inferred type.
    """
    types = {}
    for col in df.columns:
        col_data = df[col].dropna()
        n = len(col_data)
        if n == 0:
            types[col] = 'unknown'
            continue
        num_valid = pd.to_numeric(col_data, errors='coerce').notnull().sum()
        dt_valid = pd.to_datetime(col_data, errors='coerce').notnull().sum()
        if num_valid / n >= threshold:
            types[col] = 'numeric'
        elif dt_valid / n >= threshold:
            types[col] = 'datetime'
        elif num_valid / n < (1 - threshold) and dt_valid / n < (1 - threshold):
            types[col] = 'categorical'
        else:
            types[col] = 'mixed'
    return types

def rule_based_anomalies(df, types):
    """
    Detect missing, type mismatch, out-of-range, and length-inconsistent anomalies using rules.
 performance metrics of model on the dataset   Args:
        df (pd.DataFrame): Input data.
        types (dict): Column type mapping.

    Returns:
        pd.DataFrame: DataFrame of anomaly labels ('' if normal).
    """
    anomalies = pd.DataFrame('', index=df.index, columns=df.columns)
    anomalies[df.isnull()] = 'missing'
    for col in df.columns:
        if types[col] == 'numeric':
            coerced = pd.to_numeric(df[col], errors='coerce')
            mask = df[col].notnull() & coerced.isnull()
            anomalies.loc[mask, col] = 'type_mismatch'
            # Length inconsistency: flag if less than 70% match mode length
            valid_mask = (anomalies[col] == '')
            valid_values = df.loc[valid_mask, col].astype(str)
            lengths = valid_values.str.len()
            if not lengths.empty:
                mode_length = lengths.mode()[0]
                mode_count = (lengths == mode_length).sum()
                if mode_count / len(lengths) >= 0.85:
                    for idx, val in valid_values.items():
                        if len(val) != mode_length:
                            anomalies.loc[idx, col] = 'len_incon'
            # Out-of-range (IQR)
            col_data = coerced
            q1 = col_data.quantile(0.08)
            q3 = col_data.quantile(0.92)
            iqr = q3 - q1
            multiplier = 13
            lower = q1 - multiplier * iqr
            upper = q3 + multiplier * iqr
            out_range = (col_data < lower) | (col_data > upper)
            for idx in df.index:
                if anomalies.loc[idx, col] == '' and out_range.loc[idx]:
                    anomalies.loc[idx, col] = 'out_of_range'
        elif types[col] == 'datetime':
            coerced = pd.to_datetime(df[col], errors='coerce')
            mask = df[col].notnull() & coerced.isnull()
            anomalies.loc[mask, col] = 'type_mismatch'
        elif types[col] == 'mixed':
            mask = df[col].notnull()
            anomalies.loc[mask, col] = 'type_mismatch'
    return anomalies

def isolation_forest_anomalies(df, types, contamination=0.001):
    """
    Isolation Forest-based anomaly detection for numeric columns.

    Args:
        df (pd.DataFrame): Input data.
        types (dict): Column type mapping.
        contamination (float): Proportion of anomalies to expect.

    Returns:
        pd.DataFrame: DataFrame of anomaly labels ('' if normal).
    """
    anomalies = pd.DataFrame('', index=df.index, columns=df.columns)
    num_cols = [col for col in df.columns if types[col] == 'numeric']
    if num_cols and len(df) > 10:
        # Only use rows where all numeric columns are valid numbers
        valid_mask = pd.DataFrame({
            col: pd.to_numeric(df[col], errors='coerce').notnull()
            for col in num_cols
        }).all(axis=1)
        X = df.loc[valid_mask, num_cols].apply(pd.to_numeric, errors='coerce')
        if not X.empty:
            iso = IsolationForest(contamination=contamination, random_state=42)
            preds = iso.fit_predict(X)
            outlier_rows = X.index[preds == -1]
            for row in outlier_rows:
                for col in num_cols:
                    # Only mark as statistical_outlier if not already missing/type_mismatch/out_of_range
                    if anomalies.loc[row, col] == '':
                        anomalies.loc[row, col] = 'statistical_outlier'
    return anomalies

def combine_anomalies(rule_anom, iso_anom):
    """
    Combine rule-based and isolation forest anomalies, prioritizing rule-based results.

    Args:
        rule_anom (pd.DataFrame): Rule-based anomaly DataFrame.
        iso_anom (pd.DataFrame): Isolation Forest anomaly DataFrame.

    Returns:
        pd.DataFrame: Combined anomaly DataFrame.
    """
    combined = rule_anom.copy()
    for col in iso_anom.columns:
        for idx in iso_anom.index:
            if not combined.loc[idx, col] and iso_anom.loc[idx, col]:
                combined.loc[idx, col] = iso_anom.loc[idx, col]
    return combined

def replace_and_highlight(df, anomalies, output_file):
    """
    Replace anomalous cells with error label and highlight them in the Excel output.

    Args:
        df (pd.DataFrame): Original data.
        anomalies (pd.DataFrame): Anomaly labels.
        output_file (str): Path to output Excel file.
    """
    df_out = df.copy()
    for col in df.columns:
        for idx in df.index:
            if anomalies.loc[idx, col]:
                df_out.loc[idx, col] = anomalies.loc[idx, col]
    df_out.to_excel(output_file, index=False)
    wb = load_workbook(output_file)
    ws = wb.active
    fill = PatternFill(start_color=HIGHLIGHT_COLOR, end_color=HIGHLIGHT_COLOR, fill_type="solid")
    for i, col in enumerate(df.columns, 1):
        for j, idx in enumerate(df.index, 2):
            if anomalies.loc[idx, col]:
                ws.cell(row=j, column=i).fill = fill
    wb.save(output_file)

def main():
    """
    Main entry point for anomaly detection and evaluation.

    - Uses INPUT as the file name (without extension) for the model.
    - Generates an output Excel file with highlighted anomalies.
    - Prints accuracy and error analysis (missed, overpredicted, identified) by calling accuracy.py.
    - Compares highlights between input and output files by calling errors.py.
    """
    if not os.path.exists(INPUT_FILE):
        print(f"Input file not found: {INPUT_FILE}")
        return
    df = pd.read_excel(INPUT_FILE)
    types = infer_column_types(df)
    rule_anom = rule_based_anomalies(df, types)
    iso_anom = isolation_forest_anomalies(df, types)
    anomalies = combine_anomalies(rule_anom, iso_anom)
    replace_and_highlight(df, anomalies, OUTPUT_FILE)
    print("Done. Please check the output file for highlighted errors.")

if __name__ == "__main__":
    main()
    accuracy.compare_excel_highlights(INPUT_FILE, OUTPUT_FILE)
    errors.create_error_excel_combined(
        input_f=INPUT_FILE,
        output_f=OUTPUT_FILE,
        result_f=INPUT + "_missed_and_identified.xlsx"
    )
    print("Anomaly detection and comparison completed.")
